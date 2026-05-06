import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


SOURCE_DIR = Path("Financial_Statment")

STATEMENTS = {
    "Balance Sheet": [
        "consolidated statements of financial position",
        "consolidated statement of financial position",
        "consolidated balance sheets",
        "statements of financial position",
    ],
    "Income Statement": [
        "consolidated statements of profit or loss",
        "consolidated statement of profit or loss",
        "consolidated statements of income",
        "consolidated statements of operations",
        "consolidated statements of earnings",
        "consolidated results of operations",
        "profit and loss statement",
    ],
    "Cash Flow": [
        "consolidated statements of cash flows",
        "consolidated statement of cash flows",
        "statements of cash flows",
        "consolidated cash flow statements",
    ],
}

STOP_TITLES = [
    "consolidated statements of comprehensive income",
    "consolidated statement of comprehensive income",
    "consolidated statements of changes in equity",
    "consolidated statement of changes in equity",
    "notes to consolidated financial statements",
    "notes to the consolidated financial statements",
]


def clean_text(text):
    return (
        (text or "")
        .replace("\x00", "")
        .replace("\ufeff", "")
        .replace("￦", "")
        .replace("₩", "")
        .strip()
    )


def normalized(text):
    return re.sub(r"\s+", " ", clean_text(text).lower())


def extract_pages(reader, page_index):
    pages = []
    for idx in range(page_index, min(page_index + 3, len(reader.pages))):
        text = reader.pages[idx].extract_text() or ""
        lower = normalized(text)
        if idx != page_index and any(title in lower for title in STOP_TITLES):
            break
        pages.append(text)
        if "the accompanying notes are an integral part" in lower:
            break
    return "\n".join(pages)


def locate_statement_pages(reader):
    found = {}
    for idx, page in enumerate(reader.pages):
        lower = normalized(page.extract_text() or "")
        if "notes to" in lower and idx > 12:
            continue
        for sheet_name, titles in STATEMENTS.items():
            if sheet_name not in found and any(title in lower for title in titles):
                # Skip table of contents and audit-opinion references; primary statements
                # have the title plus a period/unit header near the front of the page.
                if "table of contents" in lower:
                    continue
                if "korean won in millions" in lower:
                    found[sheet_name] = idx
        if len(found) == 3:
            break
    return found


def split_header(lines):
    meta = []
    body = []
    body_started = False
    for raw in lines:
        line = clean_text(raw)
        if not line:
            continue
        if re.match(r"^-?\s*\d+\s*-?$", line):
            continue
        if line.lower().startswith("the accompanying notes are an integral part"):
            break
        if not body_started:
            meta.append(line)
            if re.search(r"\bnotes?\b", line, re.I) and (
                re.search(r"\b20\d{2}\b", line) or "december 31" in line.lower()
            ):
                body_started = True
            continue
        body.append(line)
    return meta, body


def parse_periods(meta):
    header = next((line for line in meta if re.search(r"\bnotes?\b", line, re.I)), "")
    periods = re.findall(r"December\s+31,\s*\d{4}", header, flags=re.I)
    if len(periods) < 2:
        periods = re.findall(r"\b20\d{2}\b", header)
    if len(periods) < 2:
        all_meta = " ".join(meta)
        periods = re.findall(r"\b20\d{2}\b", all_meta)
    return (periods + ["Period 1", "Period 2"])[:2]


def parse_number(value):
    value = clean_text(value)
    if value == "-":
        return "-"
    negative = value.startswith("(") and value.endswith(")")
    value = value.strip("()").replace(",", "")
    try:
        number = int(value)
    except ValueError:
        return value
    return -number if negative else number


VALUE_RE = re.compile(r"\(?-?\d[\d,]*\)?|-")
NOTE_RE = re.compile(r"^(.*?)(\d+(?:\s*,\s*\d+)*)$")


def parse_value_row(line):
    compact = re.sub(r"\s+", " ", clean_text(line))
    tokens = list(VALUE_RE.finditer(compact))
    if not tokens:
        return None

    tail = []
    cursor = len(compact)
    for match in reversed(tokens):
        between = compact[match.end() : cursor]
        if between.strip():
            break
        tail.append(match)
        cursor = match.start()
        if len(tail) == 2:
            break
    tail.reverse()
    if not tail:
        return None

    pre = compact[: tail[0].start()].strip()
    values = [parse_number(match.group()) for match in tail]
    note = ""
    label = pre
    note_match = NOTE_RE.match(pre)
    if note_match:
        possible_label, possible_note = note_match.groups()
        if possible_label.strip():
            label = possible_label.strip()
            note = re.sub(r"\s+", "", possible_note)
    return {"label": label, "note": note, "values": values}


def build_rows(text):
    raw_lines = [line.rstrip() for line in text.splitlines()]
    meta, body = split_header(raw_lines)
    periods = parse_periods(meta)
    rows = []
    pending_text = None
    pending_indent = 0

    for line in body:
        if "the accompanying notes are an integral part" in line.lower():
            break
        parsed = parse_value_row(line)
        indent = max(0, (len(line) - len(line.lstrip())) // 2)

        if parsed is None:
            if pending_text:
                rows.append([pending_text, "", "", "", pending_indent])
            pending_text = clean_text(line)
            pending_indent = indent
            continue

        label = parsed["label"]
        values = parsed["values"]
        if not label and pending_text:
            label = pending_text
            indent = pending_indent
            pending_text = None
        elif pending_text:
            rows.append([pending_text, "", "", "", pending_indent])
            pending_text = None

        if len(values) == 1:
            rows.append([label, parsed["note"], values[0], None, indent])
        else:
            rows.append([label, parsed["note"], values[0], values[1], indent])

    if pending_text:
        rows.append([pending_text, "", "", "", pending_indent])

    return meta, periods, rows


def repair_split_rows(rows):
    repaired = []
    idx = 0
    while idx < len(rows):
        row = rows[idx]
        next_row = rows[idx + 1] if idx + 1 < len(rows) else None
        if (
            next_row
            and row[2] is not None
            and row[3] is None
            and not next_row[0]
            and next_row[2] is not None
        ):
            row = [row[0], row[1], row[2], next_row[2], row[4]]
            idx += 2
            repaired.append(row)
            continue
        repaired.append(row)
        idx += 1
    return repaired


def write_sheet(wb, sheet_name, meta, periods, rows):
    ws = wb.create_sheet(sheet_name)
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    if not rows:
        ws.append(["Statement not found"])
        return

    for line in meta:
        ws.append([line])
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Line Item", "Note", periods[0], periods[1]])
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for label, note, value1, value2, indent in repair_split_rows(rows):
        ws.append([label, note, value1, value2])
        r = ws.max_row
        ws.cell(r, 1).alignment = Alignment(indent=min(indent, 8), wrap_text=True)
        if value1 == "" and value2 == "":
            ws.cell(r, 1).font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_row=min(4, ws.max_row)):
        for cell in row:
            cell.font = title_font

    for col_idx, width in enumerate([48, 14, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    for row in ws.iter_rows(min_row=header_row + 1, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, int):
                cell.number_format = '#,##0;(#,##0)'


def process_pdf(path):
    reader = PdfReader(str(path))
    pages = locate_statement_pages(reader)
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    for sheet_name in ["Income Statement", "Balance Sheet", "Cash Flow"]:
        if sheet_name not in pages:
            write_sheet(wb, sheet_name, [], [], [])
            continue
        text = extract_pages(reader, pages[sheet_name])
        meta, periods, rows = build_rows(text)
        write_sheet(wb, sheet_name, meta, periods, rows)

    output_path = path.with_suffix(".xlsx")
    wb.save(output_path)
    return output_path, pages


def main():
    for pdf_path in sorted(SOURCE_DIR.glob("*.pdf")):
        output_path, pages = process_pdf(pdf_path)
        located = ", ".join(f"{name}: page {idx + 1}" for name, idx in sorted(pages.items()))
        print(f"Wrote {output_path} ({located})")


if __name__ == "__main__":
    main()
