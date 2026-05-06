from __future__ import annotations

from openpyxl import Workbook


def build_main_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    wb.create_sheet("Income Statement")
    wb.create_sheet("Cash Flow &Comprehensive Income")
    return wb
