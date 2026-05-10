from __future__ import annotations

from financial_claw.extractor.models import ExtractionResult, StatementCandidate
from financial_claw.extractor.reconciliation import FAILED_RECONCILIATION_CELLS_KEY
from financial_claw.extractor.reconciliation import apply_reconciliation_checks


def _candidate(statement_type: str) -> StatementCandidate:
    return StatementCandidate(
        statement_type=statement_type,
        page_start=1,
        page_end=1,
        title=statement_type,
        score=1.0,
        reason="test",
        source_pages=[1],
    )


def test_balance_sheet_reconciliation_passes() -> None:
    result = ExtractionResult(
        candidate=_candidate("balance_sheet"),
        rows=[
            ["Total assets", "100"],
            ["Total liabilities", "40"],
            ["Total equity", "60"],
            ["Total liabilities and equity", "100"],
        ],
    )

    apply_reconciliation_checks([result])

    assert result.warnings == []


def test_balance_sheet_reconciliation_warns_on_mismatch() -> None:
    result = ExtractionResult(
        candidate=_candidate("balance_sheet"),
        rows=[
            ["Total assets", "102"],
            ["Total liabilities", "40"],
            ["Total equity", "60"],
        ],
    )

    apply_reconciliation_checks([result])

    assert any("Reconciliation check failed" in warning for warning in result.warnings)
    assert getattr(result, FAILED_RECONCILIATION_CELLS_KEY)


def test_income_statement_gross_profit_handles_negative_cost() -> None:
    result = ExtractionResult(
        candidate=_candidate("income_statement"),
        rows=[
            ["Revenue", "100"],
            ["Cost of sales", "(70)"],
            ["Gross profit", "30"],
        ],
    )

    apply_reconciliation_checks([result])

    assert result.warnings == []


def test_cash_flow_reconciliation_passes() -> None:
    result = ExtractionResult(
        candidate=_candidate("cash_flow"),
        rows=[
            ["Net cash provided by operating activities", "10"],
            ["Net cash used in investing activities", "(3)"],
            ["Net cash provided by financing activities", "2"],
            ["Effect of exchange rate fluctuation on cash held", "1"],
            ["Net increase in cash and cash equivalents", "10"],
            ["Cash and cash equivalents at beginning of the period", "90"],
            ["Cash and cash equivalents at end of the period", "100"],
        ],
    )

    apply_reconciliation_checks([result])

    assert result.warnings == []


def test_cash_flow_investing_section_subtotal_sums_line_items() -> None:
    result = ExtractionResult(
        candidate=_candidate("cash_flow"),
        rows=[
            ["Cash flows from investing activities", ""],
            ["Net proceeds from disposal of investment properties", "329.3"],
            ["Proceeds from disposal of controlled entities, net of cash disposed", "-"],
            ["Net proceeds from disposal of equity investments", "1.2"],
            ["Acquisitions of controlled entities, net of cash acquired", "(1,856.5)"],
            ["Payments for investment properties", "(266.6)"],
            ["Payments for investments in Partnerships", "(1,685.4)"],
            ["Payments for property, plant and equipment", "(6.4)"],
            ["Net cash used in investing activities", "(3,484.4)"],
        ],
    )

    apply_reconciliation_checks([result])

    assert result.warnings == []


def test_cash_flow_investing_section_subtotal_warns_on_bad_total() -> None:
    result = ExtractionResult(
        candidate=_candidate("cash_flow"),
        rows=[
            ["Cash flows from investing activities", ""],
            ["Net proceeds from disposal of investment properties", "329.3"],
            ["Net proceeds from disposal of equity investments", "1.2"],
            ["Acquisitions of controlled entities, net of cash acquired", "(1,856.5)"],
            ["Payments for investment properties", "(266.6)"],
            ["Payments for investments in Partnerships", "(1,685.4)"],
            ["Payments for property, plant and equipment", "(6.4)"],
            ["Net cash used in investing activities", "(3,400.0)"],
        ],
    )

    apply_reconciliation_checks([result])

    assert any("sum of investing activities line items" in warning for warning in result.warnings)


def test_cash_flow_roll_forward_requires_major_components() -> None:
    result = ExtractionResult(
        candidate=_candidate("cash_flow"),
        rows=[
            ["Cash flows from investing activities", ""],
            ["Payments for investment properties", "(100)"],
            ["Net cash used in investing activities", "(100)"],
            ["Net increase in cash and cash equivalents", "50"],
        ],
    )

    apply_reconciliation_checks([result])

    assert result.warnings == []


def test_failed_reconciliation_cells_can_be_highlighted(tmp_path) -> None:
    from openpyxl import load_workbook

    from financial_claw.extractor.excel_writer import write_workbook
    from financial_claw.extractor.models import PageProfile

    result = ExtractionResult(
        candidate=_candidate("balance_sheet"),
        rows=[
            ["Total assets", "102"],
            ["Total liabilities", "40"],
            ["Total equity", "60"],
        ],
    )
    apply_reconciliation_checks([result])

    output_path = tmp_path / "out.xlsx"
    write_workbook(
        output_path,
        {"source_pdf": "test.pdf"},
        [PageProfile(1, "", 0, 0, True, 0, 0)],
        [result.candidate],
        [result],
    )

    ws = load_workbook(output_path).active
    assert ws["B1"].fill.fgColor.rgb == "00FFC7CE"
    assert ws["B2"].fill.fgColor.rgb == "00FFC7CE"
    assert ws["B3"].fill.fgColor.rgb == "00FFC7CE"
