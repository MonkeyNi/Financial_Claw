from __future__ import annotations

from openpyxl import Workbook

from financial_claw.core.workbook_merge import parse_period_with_context, parse_statement_sheet


def test_duplicate_periods_keep_entity_context() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Consolidated statements of financial position", None, None, None, None, None])
    ws.append(["as at 30 June 2025", None, None, None, None, None])
    ws.append([None, None, "Goodman", None, "GIT", None])
    ws.append([None, None, "2025", "2024", "2025", "2024"])
    ws.append([None, "Note", "$M", "$M", "$M", "$M"])
    ws.append(["Cash and cash equivalents", None, 3957.1, 1785.3, 2470.5, 1018.2])

    parsed = parse_statement_sheet(ws)
    row = parsed.rows[0]

    assert row.values == {
        "Goodman 2025": 3957.1,
        "Goodman 2024": 1785.3,
        "GIT 2025": 2470.5,
        "GIT 2024": 1018.2,
    }


def test_duplicate_quarter_periods_do_not_use_period_descriptor_as_entity() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws.append(["Interim condensed consolidated statements of comprehensive income or loss", None, None])
    ws.append(["for each of the three-month and nine-month periods ended September 30, 2023 and 2022", None, None])
    ws.append([None, "2023", "2023"])
    ws.append(["Revenue", 100, 300])

    parsed = parse_statement_sheet(ws)

    assert parsed.rows[0].values == {"Q3 2023": 100}


def test_duration_uses_month_count_but_quarter_uses_ended_date() -> None:
    period = parse_period_with_context(
        "For the three-month period ended September 30 (unaudited) 2024",
        "",
    )

    assert period is not None
    assert period.label == "3-Month Q3 2024"


def test_duration_periods_are_kept_as_independent_columns() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws.append(
        [
            "(in millions of Won)",
            "Notes",
            "For the three-month period ended September 30 (unaudited) 2024",
            "For the nine-month period ended September 30 (unaudited) 2024",
        ]
    )
    ws.append(["Revenue", "1", 100, 900])

    parsed = parse_statement_sheet(ws)

    assert parsed.rows[0].values == {
        "3-Month Q3 2024": 100,
        "9-Month Q3 2024": 900,
    }
