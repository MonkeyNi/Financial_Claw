from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from typing import Literal

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


EXPECTED_SHEETS = ["Balance Sheet", "Income Statement", "Cash Flow Statement"]

# Excel presentation (xlsx / financial table conventions)
_FONT_FAMILY = "Arial"
_FONT_SIZE = 10
_HEADER_FILL = "1F4E78"
_HEADER_FONT_COLOR = "FFFFFF"
_ANNUAL_FILL = "D9EAF7"
_QUARTER_FILL = "E2F0D9"
_SECTION_FILL = "F2F2F2"
_TOTAL_FILL = "EAF2F8"
_BORDER_COLOR = "D9E2F3"
_STRONG_BORDER_COLOR = "7F7F7F"
_NUM_FORMAT_INT = '#,##0;(#,##0);"-"'
_NUM_FORMAT_FLOAT = '#,##0.0;(#,##0.0);"-"'


@dataclass(frozen=True)
class Period:
    label: str
    kind: Literal["annual", "quarter"]
    year: int
    quarter: int = 0
    duration_months: int = 0

    @property
    def sort_key(self) -> tuple[int, int, int]:
        return (self.year, self.quarter, self.duration_months)


@dataclass
class ParsedSheet:
    title: str
    periods: dict[int, Period]
    rows: list["StatementRow"]


@dataclass
class StatementRow:
    label: str
    key: str
    values: dict[str, object] = field(default_factory=dict)


def merge_workbooks(excel_1: Path, excel_2: Path, output_path: Path) -> Path:
    return merge_workbook_sequence([excel_1, excel_2], output_path)


def merge_workbook_sequence(workbook_paths: list[Path], output_path: Path) -> Path:
    if not workbook_paths:
        raise ValueError("At least one workbook is required for merge.")

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    parsed_workbooks = sorted(
        (_parse_workbook(path) for path in workbook_paths),
        key=lambda item: (item[0], str(item[1]).lower()),
    )
    logger.debug(
        "Merging {} workbooks in period order: {}",
        len(parsed_workbooks),
        [str(path) for _, path, _ in parsed_workbooks],
    )

    for sheet_name in EXPECTED_SHEETS:
        parsed_sheets = [workbook[sheet_name] for _, _, workbook in parsed_workbooks]
        merged_rows = list(parsed_sheets[0].rows)
        for parsed in parsed_sheets[1:]:
            logger.debug("Merging sheet: {} <- {}", sheet_name, parsed.title)
            merged_rows = merge_statement_rows(merged_rows, parsed.rows)
        periods = _merge_all_periods(parsed_sheets)
        write_merged_sheet(out_wb, sheet_name, merged_rows, periods)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    logger.debug("Merged workbook saved: {}", output_path)
    return output_path


def validate_extracted_workbook(path: Path) -> None:
    wb = load_workbook(path, data_only=True)
    _validate_workbook(wb, path)
    for sheet_name in EXPECTED_SHEETS:
        parse_statement_sheet(wb[sheet_name])


def workbook_latest_period_sort_key(path: Path) -> tuple[int, int, int]:
    return _parse_workbook(path)[0]


def _parse_workbook(path: Path) -> tuple[tuple[int, int, int], Path, dict[str, ParsedSheet]]:
    wb = load_workbook(path, data_only=True)
    _validate_workbook(wb, path)
    parsed = {sheet_name: parse_statement_sheet(wb[sheet_name]) for sheet_name in EXPECTED_SHEETS}
    periods = [period for sheet in parsed.values() for period in sheet.periods.values()]
    if not periods:
        raise ValueError(f"{path} does not contain any parseable annual or quarterly period columns.")
    latest = max(period.sort_key for period in periods)
    return latest, path, parsed


def _validate_workbook(wb, path: Path) -> None:
    names = wb.sheetnames
    if names == ["NoStatements"]:
        raise ValueError(f"{path} did not contain extracted financial statement tables; please check candidate detection or OCR output.")
    if names != EXPECTED_SHEETS:
        raise ValueError(f"{path} must contain exactly these sheets in order: {EXPECTED_SHEETS}; got {names}")


def parse_statement_sheet(ws) -> ParsedSheet:
    header_row, periods = _find_period_header(ws)
    periods = _with_entity_context_for_duplicate_periods(ws, header_row, periods)
    title = str(ws.cell(row=1, column=1).value or ws.title)
    rows: list[StatementRow] = []
    seen: dict[str, int] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if _is_unit_row(ws, row_idx, periods):
            continue
        label = _row_label(ws, row_idx, periods)
        if not label:
            continue
        key_base = normalize_item_label(label)
        if not key_base:
            continue
        seen[key_base] = seen.get(key_base, 0) + 1
        key = f"{key_base}#{seen[key_base]}"
        values: dict[str, object] = {}
        for col_idx, period in periods.items():
            # Interim statements can contain both current-quarter and YTD columns
            # that resolve to the same quarter label. Keep the first occurrence,
            # which is normally the standalone quarter column in the source table.
            values.setdefault(period.label, ws.cell(row=row_idx, column=col_idx).value)
        rows.append(StatementRow(label=label, key=key, values=values))
    return ParsedSheet(title=title, periods=periods, rows=rows)


def _find_period_header(ws) -> tuple[int, dict[int, Period]]:
    best_row = 0
    best_periods: dict[int, Period] = {}
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        context = _header_context(ws, row_idx)
        periods: dict[int, Period] = {}
        for col_idx in range(1, ws.max_column + 1):
            period = parse_period_with_context(ws.cell(row=row_idx, column=col_idx).value, context)
            if period is not None and _looks_like_period_value_column(ws, row_idx, col_idx):
                periods[col_idx] = period
        if len(periods) > len(best_periods):
            best_row = row_idx
            best_periods = periods
    if not best_periods:
        raise ValueError(f"Could not find period header row in sheet {ws.title!r}")
    return best_row, best_periods


def _with_entity_context_for_duplicate_periods(
    ws,
    header_row: int,
    periods: dict[int, Period],
) -> dict[int, Period]:
    labels = [period.label for period in periods.values()]
    if len(labels) == len(set(labels)):
        return periods

    contextual: dict[int, Period] = {}
    for col_idx, period in periods.items():
        entity = _entity_header_for_column(ws, header_row, col_idx)
        if not entity:
            contextual[col_idx] = period
            continue
        contextual[col_idx] = Period(
            label=f"{entity} {period.label}",
            kind=period.kind,
            year=period.year,
            quarter=period.quarter,
            duration_months=period.duration_months,
        )
    return contextual


def _entity_header_for_column(ws, header_row: int, col_idx: int) -> str:
    for row_idx in range(header_row - 1, 0, -1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None:
            text = ""
        else:
            text = str(value).strip()
        if _looks_like_entity_header(text):
            return text
        for left_col in range(col_idx - 1, 0, -1):
            value = ws.cell(row=row_idx, column=left_col).value
            if value is None:
                continue
            text = str(value).strip()
            if _looks_like_entity_header(text):
                return text
    return ""


def _looks_like_entity_header(text: str) -> bool:
    normalized = text.strip()
    if not normalized:
        return False
    if parse_period(normalized) is not None:
        return False
    if _looks_like_statement_title(normalized):
        return False
    if _looks_like_period_descriptor(normalized):
        return False
    if normalized.lower() in {"note", "notes", "$m", "$000", "$", "m"}:
        return False
    if _is_numeric_like(normalized):
        return False
    return bool(re.search(r"[A-Za-z]", normalized))


def _looks_like_statement_title(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", text.strip().lower())
    return bool(
        re.search(r"\bconsolidated\s+statements?\s+of\s+", normalized)
        or re.search(r"\bconsolidated\s+(?:income|cash\s+flow|balance\s+sheet)s?\b", normalized)
        or re.search(r"\binterim\s+condensed\s+consolidated\b", normalized)
    )


def _looks_like_period_descriptor(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", text.strip().lower())
    return bool(
        re.search(r"\bfor\s+(?:each\s+of\s+)?the\b", normalized)
        and re.search(r"\b(?:three|six|nine|twelve)[-\s]months?\b", normalized)
        and re.search(r"\bperiods?\b", normalized)
    )


def _is_unit_row(ws, row_idx: int, periods: dict[int, Period]) -> bool:
    period_cols = set(periods)
    values = [
        str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
        for col_idx in range(1, ws.max_column + 1)
    ]
    non_empty = [value for value in values if value]
    if not non_empty:
        return False
    allowed = {"note", "notes", "$m", "$000", "$", "m"}
    if not all(value in allowed for value in non_empty):
        return False
    return any(values[col_idx - 1] in {"$m", "$000", "$", "m"} for col_idx in period_cols)


def _looks_like_period_value_column(ws, header_row: int, col_idx: int) -> bool:
    numeric_count = 0
    text_count = 0
    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 25) + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value in (None, ""):
            continue
        if _is_numeric_like(value):
            numeric_count += 1
        else:
            text_count += 1
    return numeric_count > 0 and numeric_count >= text_count


def _is_numeric_like(value: object) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    text = str(value).strip()
    if not text:
        return False
    text = text.replace(",", "").replace(" ", "")
    if re.fullmatch(r"\(?-?\d+(?:\.\d+)?\)?", text):
        return True
    return text == "-"


def _header_context(ws, row_idx: int) -> str:
    start = max(1, row_idx - 2)
    parts: list[str] = []
    for current_row in range(start, row_idx + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=current_row, column=col_idx).value
            if value is not None:
                parts.append(str(value))
    return " ".join(parts)


def parse_period_with_context(value: object, context: str) -> Period | None:
    period = parse_period(value)
    if period is None or period.kind != "annual":
        return period

    text = str(value or "")
    if re.search(r"\b(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b", text, re.I):
        return period

    quarter = _quarter_from_context(context)
    if quarter is None:
        return period
    duration_months = _duration_months_from_context(context)
    return Period(
        label=_quarter_label(period.year, quarter, duration_months),
        kind="quarter",
        year=period.year,
        quarter=quarter,
        duration_months=duration_months,
    )


def parse_period(value: object) -> Period | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = re.sub(r"\s+", " ", text.replace("\n", " "))

    quarter_match = re.search(r"\b(?:(\d{1,2})\s*-\s*months?\s+)?Q([1-4])\s*(20\d{2})\b", normalized, re.I)
    if not quarter_match:
        quarter_match = re.search(r"\b(20\d{2})\s*Q([1-4])\b", normalized, re.I)
        if quarter_match:
            year = int(quarter_match.group(1))
            quarter = int(quarter_match.group(2))
            return Period(label=_quarter_label(year, quarter, 0), kind="quarter", year=year, quarter=quarter)
    else:
        duration_months = int(quarter_match.group(1) or 0)
        quarter = int(quarter_match.group(2))
        year = int(quarter_match.group(3))
        return Period(
            label=_quarter_label(year, quarter, duration_months),
            kind="quarter",
            year=year,
            quarter=quarter,
            duration_months=duration_months,
        )

    fy_quarter_match = re.search(r"\bFY(\d{2})\s*([1-4])Q\b", normalized, re.I)
    if fy_quarter_match:
        year = 2000 + int(fy_quarter_match.group(1))
        quarter = int(fy_quarter_match.group(2))
        return Period(label=_quarter_label(year, quarter, 0), kind="quarter", year=year, quarter=quarter)

    interim_period = _parse_interim_period(normalized)
    if interim_period is not None:
        return interim_period

    year_match = re.search(r"\b(20\d{2})\b", normalized)
    if year_match:
        year = int(year_match.group(1))
        return Period(label=str(year), kind="annual", year=year)
    return None


def _parse_interim_period(text: str) -> Period | None:
    normalized = re.sub(r"\s+", " ", text.replace("(", " (")).strip()
    month_day_year = re.search(
        r"\b(march|jun(?:e)?|sept(?:ember)?)\s+([0-3]?\d)\b[^0-9]{0,80}\b(20\d{2})\b",
        normalized,
        re.I,
    )
    if not month_day_year:
        return None

    month = month_day_year.group(1).lower()
    day = int(month_day_year.group(2))
    year = int(month_day_year.group(3))
    quarter = _quarter_from_month_day(month, day)
    if quarter is None:
        return None
    duration_months = _duration_months_from_context(normalized)
    return Period(
        label=_quarter_label(year, quarter, duration_months),
        kind="quarter",
        year=year,
        quarter=quarter,
        duration_months=duration_months,
    )


def _quarter_from_context(text: str) -> int | None:
    normalized = re.sub(r"\s+", " ", text).lower()
    if re.search(r"\bsept?(?:ember)?\s+30\b", normalized):
        return 3
    if re.search(r"\bjun(?:e)?\s+30\b", normalized):
        return 2
    if re.search(r"\bmar(?:ch)?\s+31\b", normalized):
        return 1
    return None


def _duration_months_from_context(text: str) -> int:
    normalized = re.sub(r"\s+", " ", text).lower()
    durations: set[int] = set()
    for word, months in (("three", 3), ("six", 6), ("nine", 9), ("twelve", 12)):
        if re.search(rf"\b{word}[-\s]months?\b", normalized):
            durations.add(months)
    for match in re.finditer(r"\b(3|6|9|12)[-\s]months?\b", normalized):
        durations.add(int(match.group(1)))
    return next(iter(durations)) if len(durations) == 1 else 0


def _quarter_label(year: int, quarter: int, duration_months: int) -> str:
    base = f"Q{quarter} {year}"
    return f"{duration_months}-Month {base}" if duration_months else base


def _quarter_from_month_day(month: str, day: int) -> int | None:
    if month.startswith("mar") and day == 31:
        return 1
    if month.startswith("jun") and day == 30:
        return 2
    if month.startswith("sep") and day == 30:
        return 3
    return None


def _row_label(ws, row_idx: int, periods: dict[int, Period]) -> str:
    period_cols = set(periods)
    label_parts: list[str] = []
    first_period_col = min(period_cols)
    for col_idx in range(1, first_period_col):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None:
            continue
        text = str(value).strip()
        if not text or _is_note_cell(text):
            continue
        label_parts.append(text)
    return " ".join(label_parts).strip()


def _is_note_cell(text: str) -> bool:
    normalized = text.strip()
    return bool(re.fullmatch(r"\d+(?:\s*,\s*\d+)*(?:\s*\([a-z]\))?", normalized, re.I))


def normalize_item_label(label: str) -> str:
    text = label.lower()
    text = text.replace("&", "and")
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\b(total|net)\b", r" \1 ", text)
    return re.sub(r"\s+", " ", text).strip()


def merge_statement_rows(rows_1: list[StatementRow], rows_2: list[StatementRow]) -> list[StatementRow]:
    merged: list[StatementRow] = []
    used_right: set[int] = set()
    right_keys = {row.key: idx for idx, row in enumerate(rows_2)}

    for left in rows_1:
        right_idx = right_keys.get(left.key)
        if right_idx is None:
            right_idx = _find_conservative_match(left, rows_2, used_right)
        if right_idx is None:
            merged.append(StatementRow(label=left.label, key=left.key, values=dict(left.values)))
            continue

        right = rows_2[right_idx]
        used_right.add(right_idx)
        values = dict(left.values)
        values.update(right.values)
        merged.append(StatementRow(label=right.label, key=left.key, values=values))

    for idx, right in enumerate(rows_2):
        if idx in used_right:
            continue
        merged.append(StatementRow(label=right.label, key=right.key, values=dict(right.values)))
    return merged


def _find_conservative_match(
    left: StatementRow,
    rows_2: list[StatementRow],
    used_right: set[int],
) -> int | None:
    best_idx: int | None = None
    best_score = 0.0
    left_base = left.key.rsplit("#", 1)[0]
    for idx, right in enumerate(rows_2):
        if idx in used_right:
            continue
        right_base = right.key.rsplit("#", 1)[0]
        score = SequenceMatcher(None, left_base, right_base).ratio()
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx if best_score >= 0.96 else None


def _merge_periods(left: ParsedSheet, right: ParsedSheet) -> list[Period]:
    return _merge_all_periods([left, right])


def _merge_all_periods(parsed_sheets: list[ParsedSheet]) -> list[Period]:
    by_label: dict[str, Period] = {}
    for parsed in parsed_sheets:
        for period in parsed.periods.values():
            by_label[period.label] = period
    annual = sorted((p for p in by_label.values() if p.kind == "annual"), key=lambda p: p.sort_key)
    quarter = sorted((p for p in by_label.values() if p.kind == "quarter"), key=lambda p: p.sort_key)
    return annual + quarter


def write_merged_sheet(wb: Workbook, sheet_name: str, rows: list[StatementRow], periods: list[Period]) -> None:
    ws = wb.create_sheet(sheet_name)
    annual = [p for p in periods if p.kind == "annual"]
    quarter = [p for p in periods if p.kind == "quarter"]

    headers = ["Item"] + [p.label for p in annual]
    blank_col = None
    if quarter:
        blank_col = len(headers) + 1
        headers.append("")
        headers.extend(p.label for p in quarter)
    ws.append(headers)

    ordered_periods = annual + quarter
    for row in rows:
        values = [row.label]
        values.extend(row.values.get(p.label, "") for p in annual)
        if quarter:
            values.append("")
            values.extend(row.values.get(p.label, "") for p in quarter)
        ws.append(values)

    _style_sheet(ws, blank_col)


def _is_numeric_cell_value(value: object) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float, Decimal))


def _number_format_for_value(value: int | float | Decimal) -> str:
    if isinstance(value, float) and not value.is_integer():
        return _NUM_FORMAT_FLOAT
    if isinstance(value, Decimal) and (value % 1) != 0:
        return _NUM_FORMAT_FLOAT
    return _NUM_FORMAT_INT


def _style_sheet(ws, blank_col: int | None) -> None:
    thin = Side(style="thin", color=_BORDER_COLOR)
    strong = Side(style="thin", color=_STRONG_BORDER_COLOR)
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    table_right_border = Border(left=thin, right=strong, top=thin, bottom=thin)
    total_border = Border(left=thin, right=thin, top=strong, bottom=thin)
    total_right_border = Border(left=thin, right=strong, top=strong, bottom=thin)
    header_right_border = Border(left=thin, right=strong, top=thin, bottom=thin)
    header_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color=_HEADER_FONT_COLOR)
    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
    body_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, color="000000")
    body_numeric_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, color="000000")
    section_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
    total_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
    total_numeric_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
    section_fill = PatternFill(fill_type="solid", fgColor=_SECTION_FILL)
    total_fill = PatternFill(fill_type="solid", fgColor=_TOTAL_FILL)
    annual_fill = PatternFill(fill_type="solid", fgColor=_ANNUAL_FILL)
    quarter_fill = PatternFill(fill_type="solid", fgColor=_QUARTER_FILL)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_right_border if col_idx == ws.max_column else grid_border
        if col_idx != 1 and (not blank_col or col_idx != blank_col):
            cell.fill = quarter_fill if blank_col and col_idx > blank_col else annual_fill
            cell.font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
        if blank_col and col_idx == blank_col:
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    max_row = ws.max_row
    max_col = ws.max_column
    for row_idx in range(2, max_row + 1):
        row_label = str(ws.cell(row=row_idx, column=1).value or "").strip()
        is_section = _is_section_row(ws, row_idx, blank_col)
        is_total = _is_total_row_label(row_label)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = table_right_border if col_idx == max_col else grid_border
            val = cell.value
            if is_section:
                cell.fill = section_fill
                cell.font = section_font
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="top", wrap_text=True)
                continue
            if is_total:
                cell.fill = total_fill
                cell.border = total_right_border if col_idx == max_col else total_border
            if col_idx == 1:
                cell.font = total_font if is_total else body_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                continue
            if blank_col and col_idx == blank_col:
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                continue
            if _is_numeric_cell_value(val):
                cell.font = total_numeric_font if is_total else body_numeric_font
                cell.number_format = _number_format_for_value(val)
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.font = total_font if is_total else body_font
                cell.alignment = Alignment(horizontal="right", vertical="top")

    if blank_col:
        ws.column_dimensions[ws.cell(row=1, column=blank_col).column_letter].width = 4
    for col_idx in range(1, ws.max_column + 1):
        if blank_col and col_idx == blank_col:
            continue
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            parts = str(value).splitlines() or [""]
            max_len = max(max_len, max(len(part) for part in parts))
        letter = ws.cell(row=1, column=col_idx).column_letter
        if col_idx == 1:
            ws.column_dimensions[letter].width = max(36, min(70, max_len + 2))
        else:
            ws.column_dimensions[letter].width = max(14, min(18, max_len + 2))


def _is_section_row(ws, row_idx: int, blank_col: int | None) -> bool:
    values = [
        ws.cell(row=row_idx, column=col_idx).value
        for col_idx in range(2, ws.max_column + 1)
        if not blank_col or col_idx != blank_col
    ]
    return bool(ws.cell(row=row_idx, column=1).value) and all(value in (None, "") for value in values)


def _is_total_row_label(label: str) -> bool:
    normalized = label.strip().lower()
    if not normalized:
        return False
    return normalized.startswith(
        (
            "total ",
            "net ",
            "gross profit",
            "operating profit",
            "profit for",
            "profit before",
            "cash and cash equivalents at end",
        )
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge one extracted financial statement workbook into another.")
    parser.add_argument("excel_1", help="Base workbook.")
    parser.add_argument("excel_2", help="Workbook to merge into the base workbook.")
    parser.add_argument("-o", "--output", required=True, help="Merged output workbook path.")
    args = parser.parse_args()

    merge_workbooks(Path(args.excel_1), Path(args.excel_2), Path(args.output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
