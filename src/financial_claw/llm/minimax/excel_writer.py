from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from financial_claw.llm.minimax.table_schema import TableCell, TableDocument, TableSheet


def _hex_to_argb(hex_color: str) -> str | None:
    s = (hex_color or "").strip()
    if not s:
        return None
    if s.startswith("#"):
        s = s[1:]
    if len(s) == 6:
        return "FF" + s.upper()
    if len(s) == 8:
        return s.upper()
    return None


def _apply_cell_style(xl_cell, cell: TableCell) -> None:
    st = cell.style

    if st.bold is not None or st.italic is not None or st.font_color:
        xl_cell.font = Font(
            bold=bool(st.bold) if st.bold is not None else None,
            italic=bool(st.italic) if st.italic is not None else None,
            color=_hex_to_argb(st.font_color) if st.font_color else None,
        )

    if st.align or st.valign:
        horiz = None if st.align in (None, "general") else st.align
        vert = None if st.valign in (None, "general") else st.valign
        xl_cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=True)

    if st.number_format:
        xl_cell.number_format = st.number_format

    if st.bg_color:
        argb = _hex_to_argb(st.bg_color)
        if argb:
            xl_cell.fill = PatternFill("solid", fgColor=argb)


def write_table_document_to_xlsx(doc: TableDocument, out_path: Path) -> None:
    wb = Workbook()
    if not doc.sheets:
        doc = TableDocument(
            sheets=[TableSheet(name="Sheet1", cells=[TableCell(r=1, c=1, v="(empty)")])]
        )

    # Replace default sheet if present
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet in doc.sheets:
        ws = wb.create_sheet(title=sheet.name[:31] or "Sheet1")

        # Set widths/heights (optional)
        if sheet.column_widths:
            for idx, w in enumerate(sheet.column_widths, start=1):
                if w and w > 0:
                    ws.column_dimensions[get_column_letter(idx)].width = float(w)
        if sheet.row_heights:
            for idx, h in enumerate(sheet.row_heights, start=1):
                if h and h > 0:
                    ws.row_dimensions[idx].height = float(h)

        # Write cells
        merges: list[tuple[int, int, int, int]] = []
        for cell in sheet.cells:
            if cell.r <= 0 or cell.c <= 0:
                continue
            xl_cell = ws.cell(row=cell.r, column=cell.c, value=cell.v)
            _apply_cell_style(xl_cell, cell)
            if cell.rowspan > 1 or cell.colspan > 1:
                merges.append((cell.r, cell.c, cell.r + cell.rowspan - 1, cell.c + cell.colspan - 1))

        for r1, c1, r2, c2 in merges:
            try:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            except Exception as e:
                logger.warning("Failed to merge {}:{} -> {}:{}: {}", r1, c1, r2, c2, e)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info("Saved xlsx: {}", str(out_path))
    if doc.meta:
        logger.debug("Document meta: {}", asdict(doc).get("meta"))
