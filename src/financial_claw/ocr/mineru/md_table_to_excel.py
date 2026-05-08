#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from loguru import logger


@dataclass(frozen=True)
class _Cell:
    text: str
    rowspan: int = 1
    colspan: int = 1


class _HtmlTablesParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._in_table = False
        self._in_tr = False
        self._in_cell = False
        self._cell_tag: Optional[str] = None
        self._cell_attrs: Dict[str, str] = {}
        self._cell_text_parts: List[str] = []

        self._current_rows: List[List[_Cell]] = []
        self.tables: List[List[List[_Cell]]] = []

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        tag = tag.lower()
        attrs_dict: Dict[str, str] = {k.lower(): (v or "") for k, v in attrs}

        if tag == "table":
            self._in_table = True
            self._current_rows = []
            return

        if not self._in_table:
            return

        if tag == "tr":
            self._in_tr = True
            self._current_rows.append([])
            return

        if self._in_tr and tag in ("td", "th"):
            self._in_cell = True
            self._cell_tag = tag
            self._cell_attrs = attrs_dict
            self._cell_text_parts = []
            return

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()

        if tag == "table":
            if self._in_table:
                self.tables.append([row[:] for row in self._current_rows])
            self._in_table = False
            self._in_tr = False
            self._in_cell = False
            self._cell_tag = None
            self._cell_attrs = {}
            self._cell_text_parts = []
            self._current_rows = []
            return

        if not self._in_table:
            return

        if tag == "tr":
            self._in_tr = False
            return

        if self._in_cell and tag == self._cell_tag:
            text = _normalize_cell_text("".join(self._cell_text_parts))
            rowspan = _parse_span(self._cell_attrs.get("rowspan"))
            colspan = _parse_span(self._cell_attrs.get("colspan"))
            if not self._current_rows:
                self._current_rows.append([])
            self._current_rows[-1].append(_Cell(text=text, rowspan=rowspan, colspan=colspan))

            self._in_cell = False
            self._cell_tag = None
            self._cell_attrs = {}
            self._cell_text_parts = []
            return

    def handle_data(self, data: str) -> None:
        if self._in_table and self._in_cell:
            self._cell_text_parts.append(data)

    def error(self, message: str) -> None:  # pragma: no cover
        raise ValueError(message)


def _parse_span(v: Optional[str]) -> int:
    if not v:
        return 1
    try:
        n = int(v.strip())
        return n if n >= 1 else 1
    except ValueError:
        return 1


_WS_RE = re.compile(r"\s+")


def _normalize_cell_text(s: str) -> str:
    s = s.replace("\xa0", " ")
    s = _WS_RE.sub(" ", s).strip()
    return s


def _table_cells_to_grid(rows: List[List[_Cell]]) -> List[List[str]]:
    """
    Expand a table represented by rows of cells into a full 2D grid of strings,
    honoring rowspan/colspan by expanding the occupied area.

    Note: For usability in Excel, we only place the text in the top-left cell of
    the spanned region, leaving the rest blank (instead of duplicating text).
    """
    grid: List[List[Optional[str]]] = []
    occupied: List[List[bool]] = []

    def ensure_size(r: int, c: int) -> None:
        while len(grid) <= r:
            grid.append([])
            occupied.append([])
        while len(grid[r]) <= c:
            grid[r].append(None)
            occupied[r].append(False)

    def next_free_col(r: int, start_c: int) -> int:
        c = start_c
        while True:
            ensure_size(r, c)
            if not occupied[r][c]:
                return c
            c += 1

    for r, row in enumerate(rows):
        c = 0
        for cell in row:
            c = next_free_col(r, c)

            rs = max(1, cell.rowspan)
            cs = max(1, cell.colspan)
            for rr in range(r, r + rs):
                for cc in range(c, c + cs):
                    ensure_size(rr, cc)
                    grid[rr][cc] = cell.text if (rr == r and cc == c) else ""
                    occupied[rr][cc] = True
            c += cs

    # Make rectangular
    max_cols = max((len(r) for r in grid), default=0)
    rect: List[List[str]] = []
    for r in grid:
        rr = [(v if v is not None else "") for v in r]
        if len(rr) < max_cols:
            rr.extend([""] * (max_cols - len(rr)))
        rect.append(rr)
    return rect


def _extract_tables(markdown_text: str) -> List[List[List[_Cell]]]:
    parser = _HtmlTablesParser()
    parser.feed(markdown_text)
    parser.close()
    return parser.tables


def _write_xlsx(grids: List[List[List[str]]], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install it with: pip install openpyxl"
        ) from e

    wb = Workbook()
    # Remove default empty sheet if we have our own content.
    default_ws = wb.active
    if grids:
        wb.remove(default_ws)

    for i, grid in enumerate(grids, start=1):
        ws = wb.create_sheet(title=f"Table{i}")
        for r_idx, row in enumerate(grid, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Basic usability: freeze header row if it looks like one.
        if grid and any(v for v in grid[0]):
            ws.freeze_panes = "A2"

        # Light auto-width (cheap heuristic).
        col_widths: Dict[int, int] = {}
        for row in grid[:200]:  # cap work for huge tables
            for c_idx, value in enumerate(row, start=1):
                col_widths[c_idx] = max(col_widths.get(c_idx, 0), len(str(value)))
        for c_idx, w in col_widths.items():
            ws.column_dimensions[_xlsx_col_letter(c_idx)].width = min(max(w + 2, 10), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        alt = _next_available_path(output_path)
        logger.warning("No permission to overwrite. Saving to: {}", alt)
        wb.save(alt)


def _next_available_path(path: Path) -> Path:
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for i in range(1, 10_000):
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not find available filename near {path}")


def _xlsx_col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(ord("A") + rem) + s
    return s


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Convert Markdown-embedded HTML <table> to Excel (.xlsx).")
    ap.add_argument(
        "input",
        nargs="?",
        default=str(Path(__file__).with_name("tmp.md")),
        help="Input markdown path (default: tmp.md next to this script)",
    )
    ap.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output xlsx path (default: <input_basename>.xlsx in the same directory)",
    )
    args = ap.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        logger.error("Input file not found: {}", input_path)
        return 2

    output_path = Path(args.output).expanduser().resolve() if args.output else input_path.with_suffix(".xlsx")

    logger.info("Reading markdown: {}", input_path)
    text = input_path.read_text(encoding="utf-8", errors="replace")

    tables = _extract_tables(text)
    if not tables:
        logger.error("No <table> found in markdown: {}", input_path)
        return 3

    grids = [_table_cells_to_grid(t) for t in tables]
    logger.info("Found {} table(s). Writing xlsx: {}", len(grids), output_path)
    _write_xlsx(grids, output_path)
    logger.info("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
