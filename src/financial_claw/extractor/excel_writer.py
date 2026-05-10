from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .cleaning_config import CELL_TEXT_REPLACEMENTS
from .models import ExtractionResult, PageProfile, StatementCandidate
from .reconciliation import FAILED_RECONCILIATION_CELLS_KEY

_FONT_FAMILY = "Arial"
_FONT_SIZE = 10
_HEADER_FILL = "1F4E78"
_HEADER_FONT_COLOR = "FFFFFF"
_ANNUAL_FILL = "D9EAF7"
_SECTION_FILL = "F2F2F2"
_TOTAL_FILL = "EAF2F8"
_BORDER_COLOR = "D9E2F3"
_STRONG_BORDER_COLOR = "7F7F7F"
DEFAULT_FONT = Font(name=_FONT_FAMILY, size=_FONT_SIZE, color="000000")
HEADER_FONT = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color=_HEADER_FONT_COLOR)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
YELLOW_ASSUMPTION_FILL = PatternFill(fill_type="solid", fgColor="FFF200")
RECONCILIATION_FAILURE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
INTEGER_FORMAT = '#,##0;(#,##0);"-"'
DECIMAL_FORMAT = '#,##0.######;(#,##0.######);"-"'
SCORE_FORMAT = '0.0000;(-0.0000);"-"'

NUMBER_TOKEN_RE = re.compile(r"^\(?-?(?:\d+|\d{1,3}(?:,\d{3})+)(?:\.\d+)?\)?$|^-$")
MILLION = Decimal("1000000")
THOUSAND_TO_MILLION = Decimal("0.001")
BILLION_TO_MILLION = Decimal("1000")


def write_workbook(
    output_path: Path,
    metadata: dict[str, str],
    profiles: list[PageProfile],
    candidates: list[StatementCandidate],
    results: list[ExtractionResult],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    records: list[dict[str, str]] = []

    used_names: set[str] = set()
    type_counts = _statement_type_counts(results)
    type_seen: dict[str, int] = {}
    for idx, result in enumerate(results, start=1):
        statement_type = result.candidate.statement_type
        type_seen[statement_type] = type_seen.get(statement_type, 0) + 1
        name = _sheet_name(result.candidate, type_seen[statement_type], type_counts[statement_type], used_names)
        used_names.add(name)
        out_ws = wb.create_sheet(name)
        note_columns = _note_columns(result.rows)
        million_scale = _million_scale_for_rows(result.rows)
        failed_reconciliation_cells = set(getattr(result, FAILED_RECONCILIATION_CELLS_KEY, []))
        for row in result.rows:
            convert_monetary_values = not _is_non_monetary_row(row)
            out_ws.append(
                [
                    _coerce_excel_cell_value(
                        cell,
                        force_text=(col_idx in note_columns),
                        million_scale=million_scale if convert_monetary_values else Decimal("1"),
                        normalize_unit_text=million_scale != Decimal("1"),
                    )
                    for col_idx, cell in enumerate(row)
                ]
            )
        _highlight_failed_reconciliation_cells(out_ws, failed_reconciliation_cells)
        _style_statement_sheet(out_ws)
        _highlight_failed_reconciliation_cells(out_ws, failed_reconciliation_cells)
        records.append(_record_row(output_path, metadata, result, name))
    if not wb.sheetnames:
        no_statement_ws = wb.create_sheet("NoStatements")
        no_statement_ws["A1"] = "No statements extracted."
        _style_statement_sheet(no_statement_ws)
    wb.save(output_path)
    write_extraction_record(output_path.with_name(f"{output_path.stem}_extraction_record.xlsx"), records)


def write_extraction_record(output_path: Path, records: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "extraction_record"
    headers = [
        "output_workbook",
        "sheet_name",
        "statement_type",
        "source_pdf",
        "source_pages",
        "page_start",
        "page_end",
        "extraction_method",
        "title",
        "score",
        "warnings",
    ]
    ws.append(headers)
    for record in records:
        ws.append([_coerce_excel_cell_value(record.get(header, "")) for header in headers])
    _style_record_sheet(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _record_row(
    output_workbook: Path,
    metadata: dict[str, str],
    result: ExtractionResult,
    sheet_name: str,
) -> dict[str, str]:
    candidate = result.candidate
    return {
        "output_workbook": str(output_workbook),
        "sheet_name": sheet_name,
        "statement_type": candidate.statement_type,
        "source_pdf": metadata.get("source_pdf", ""),
        "source_pages": ", ".join(str(page) for page in candidate.source_pages),
        "page_start": str(candidate.page_start),
        "page_end": str(candidate.page_end),
        "extraction_method": candidate.extraction_method,
        "title": candidate.title,
        "score": str(candidate.score),
        "warnings": "; ".join(result.warnings),
    }


def _statement_type_counts(results: list[ExtractionResult]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for result in results:
        statement_type = result.candidate.statement_type
        counts[statement_type] = counts.get(statement_type, 0) + 1
    return counts


def _sheet_name(candidate: StatementCandidate, sequence: int, total_for_type: int, used_names: set[str]) -> str:
    base = {
        "balance_sheet": "Balance Sheet",
        "income_statement": "Income Statement",
        "cash_flow": "Cash Flow Statement",
    }.get(candidate.statement_type, candidate.statement_type)
    title = f"{base} {sequence:02d}" if total_for_type > 1 else base
    title = re.sub(r"[\[\]:*?/\\]", "_", title)[:31]
    while title in used_names:
        sequence += 1
        suffix = f" {sequence:02d}"
        title = f"{base[:31-len(suffix)]}{suffix}"
    return title


def _note_columns(rows: list[list[str]]) -> set[int]:
    note_columns: set[int] = set()
    for row in rows[:8]:
        for idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() in {"note", "notes"}:
                note_columns.add(idx)
    return note_columns


def _coerce_excel_cell_value(
    value: object,
    force_text: bool = False,
    million_scale: Decimal = Decimal("1"),
    normalize_unit_text: bool = False,
) -> object:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value

    raw_text = value.strip()
    if re.fullmatch(r"[¥￦]\s*\d{1,2}", raw_text):
        return raw_text

    text = _apply_cell_text_replacements(value).strip()
    if not text:
        return ""
    if normalize_unit_text:
        text = _normalize_unit_text_to_millions(text)
    if force_text:
        return text
    if re.fullmatch(r"\d{4}", text):
        return text
    normalized_number = _normalize_number_text(text)
    if not NUMBER_TOKEN_RE.match(normalized_number):
        return text
    if text in {"-", "–"}:
        return text

    negative = normalized_number.startswith("(") and normalized_number.endswith(")")
    cleaned = normalized_number.strip("()").replace(",", "")
    try:
        number = Decimal(cleaned) * million_scale
    except (InvalidOperation, ValueError):
        return text
    if negative:
        number = -number
    return int(number) if number == number.to_integral_value() else number


def _million_scale_for_rows(rows: list[list[str]]) -> Decimal:
    text = " ".join(
        str(cell)
        for row in rows[:12]
        for cell in row
        if str(cell or "").strip()
    ).lower()
    text = re.sub(r"\s+", " ", text)
    if not text:
        return Decimal("1")
    if re.search(r"(?<![a-z0-9])(?:\$|us\$|hk\$|krw|won|w|₩)\s*m\b", text):
        return Decimal("1")
    if re.search(r"(?<![a-z0-9])(?:\$|us\$|hk\$|krw|won|w|₩)\s*0{3}\b", text) or re.search(r"\$000\b", text):
        return THOUSAND_TO_MILLION
    if re.search(r"\b(?:in|expressed in|amounts? in)[^.;:()]{0,80}\bthousands?\b", text):
        return THOUSAND_TO_MILLION
    if re.search(r"\b(?:in|expressed in|amounts? in)[^.;:()]{0,80}\bbillions?\b", text):
        return BILLION_TO_MILLION
    if re.search(r"\b(?:in|expressed in|amounts? in)[^.;:()]{0,80}\bmillions?\b", text):
        return Decimal("1")
    if re.search(r"\b(?:in|expressed in|amounts? in)[^.;:()]{0,80}\b(?:won|dollars?|usd|krw)\b", text):
        return Decimal("1") / MILLION
    return Decimal("1")


def _is_non_monetary_row(row: list[str]) -> bool:
    label = str(row[0] if row else "").strip().lower()
    return bool(
        re.search(r"\bper share\b", label)
        or re.search(r"\bearnings? per\b", label)
        or re.search(r"\bnumber of shares?\b", label)
        or re.search(r"\bweighted average\b", label)
    )


def _normalize_unit_text_to_millions(text: str) -> str:
    normalized = re.sub(r"\bin\s+thousands?\b", "in millions", text, flags=re.I)
    normalized = re.sub(r"\bin\s+billions?\b", "in millions", normalized, flags=re.I)
    normalized = re.sub(
        r"\b(expressed)\s+in\s+((?:korean\s+)?won|krw|(?:u\.s\.\s+)?dollars?|usd)\b",
        lambda match: f"{match.group(1)} in millions of {match.group(2)}",
        normalized,
        flags=re.I,
    )
    normalized = re.sub(
        r"\b(amounts?)\s+in\s+((?:korean\s+)?won|krw|(?:u\.s\.\s+)?dollars?|usd)\b",
        lambda match: f"{match.group(1)} in millions of {match.group(2)}",
        normalized,
        flags=re.I,
    )
    normalized = re.sub(r"\$000\b", "$m", normalized, flags=re.I)
    return normalized


def _apply_cell_text_replacements(text: str) -> str:
    for old, new in CELL_TEXT_REPLACEMENTS.items():
        text = text.replace(old, new)
    return text


def _normalize_number_text(text: str) -> str:
    cleaned = _strip_currency_prefix(text)
    if cleaned.startswith("W "):
        cleaned = cleaned[2:].strip()
    if cleaned.endswith(" W"):
        cleaned = cleaned[:-2].strip()
    return cleaned


def _strip_currency_prefix(text: str) -> str:
    cleaned = text.strip()
    cleaned = re.sub(r"^(?:US\$|HK\$|KRW|USD|\$|￦|₩|¥|鈧\?|锟\?|�)\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(r"\s*(?:US\$|HK\$|KRW|USD|\$|￦|₩|¥|鈧\?|锟\?|�)$", "", cleaned, flags=re.I)
    return cleaned.strip()


def _style_statement_sheet(ws) -> None:
    if ws.max_row == 0 or ws.max_column == 0:
        return
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    _apply_financial_table_style(ws, statement_sheet=True)


def _style_record_sheet(ws) -> None:
    if ws.max_row == 0:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _apply_financial_table_style(ws, statement_sheet=False)
    _highlight_attention_columns(ws)


def _highlight_failed_reconciliation_cells(ws, cells: set[tuple[int, int]]) -> None:
    for row_idx, col_idx in cells:
        ws.cell(row=row_idx + 1, column=col_idx + 1).fill = RECONCILIATION_FAILURE_FILL


def _is_reconciliation_failure_cell(cell) -> bool:
    return cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb == "00FFC7CE"


def _apply_financial_table_style(ws, *, statement_sheet: bool) -> None:
    thin = Side(style="thin", color=_BORDER_COLOR)
    strong = Side(style="thin", color=_STRONG_BORDER_COLOR)
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    table_right_border = Border(left=thin, right=strong, top=thin, bottom=thin)
    total_border = Border(left=thin, right=thin, top=strong, bottom=thin)
    total_right_border = Border(left=thin, right=strong, top=strong, bottom=thin)
    header_right_border = Border(left=thin, right=strong, top=thin, bottom=thin)
    body_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, color="000000")
    section_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
    total_font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")
    section_fill = PatternFill(fill_type="solid", fgColor=_SECTION_FILL)
    total_fill = PatternFill(fill_type="solid", fgColor=_TOTAL_FILL)
    period_fill = PatternFill(fill_type="solid", fgColor=_ANNUAL_FILL)

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22
    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx in range(1, max_row + 1):
        row_label = str(ws.cell(row=row_idx, column=1).value or "").strip()
        is_first_row = row_idx == 1
        is_section = statement_sheet and _is_section_row(ws, row_idx)
        is_period = statement_sheet and _is_period_header_row(ws, row_idx)
        is_total = statement_sheet and _is_total_row_label(row_label)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            cell.border = table_right_border if col_idx == max_col else grid_border
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "right",
                vertical="top",
                wrap_text=True,
            )

            if is_first_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center", wrap_text=True)
                cell.border = header_right_border if col_idx == max_col else grid_border
                continue

            if _is_reconciliation_failure_cell(cell):
                pass
            elif is_period:
                cell.fill = period_fill
                cell.font = Font(name=_FONT_FAMILY, size=_FONT_SIZE, bold=True, color="000000")

            if _is_reconciliation_failure_cell(cell):
                pass
            elif is_section:
                cell.fill = section_fill
                cell.font = section_font

            if _is_reconciliation_failure_cell(cell):
                pass
            elif is_total:
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_right_border if col_idx == max_col else total_border

            if _is_numeric_cell_value(val):
                cell.number_format = SCORE_FORMAT if _is_score_column(ws, col_idx) else _number_format_for_value(val)
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    _autofit_columns(ws)


def _highlight_attention_columns(ws) -> None:
    headers = [str(cell.value or "") for cell in ws[1]]
    for key in ("warnings",):
        if key not in headers:
            continue
        col = headers.index(key) + 1
        ws.cell(row=1, column=col).fill = YELLOW_ASSUMPTION_FILL


def _is_numeric_cell_value(value: object) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float, Decimal))


def _number_format_for_value(value: int | float | Decimal) -> str:
    if isinstance(value, float) and not value.is_integer():
        return DECIMAL_FORMAT
    if isinstance(value, Decimal) and value != value.to_integral_value():
        return DECIMAL_FORMAT
    return INTEGER_FORMAT


def _is_score_column(ws, col_idx: int) -> bool:
    score_col = _find_column_by_header(ws, "score")
    return score_col == col_idx


def _find_column_by_header(ws, header_name: str) -> int | None:
    if ws.max_row < 1:
        return None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip().lower() == header_name:
            return idx
    return None


def _is_section_row(ws, row_idx: int) -> bool:
    values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(2, ws.max_column + 1)]
    return bool(ws.cell(row=row_idx, column=1).value) and all(value in (None, "") for value in values)


def _is_period_header_row(ws, row_idx: int) -> bool:
    row_values = [str(ws.cell(row=row_idx, column=col_idx).value or "").strip() for col_idx in range(1, ws.max_column + 1)]
    text = " ".join(value for value in row_values if value).lower()
    if not text:
        return False
    if "note" in text and re.search(r"\b(?:19|20)\d{2}\b", text):
        return True
    if "for the " in text and re.search(r"\b(?:19|20)\d{2}\b", text):
        return True
    if "december 31" in text or "march 31" in text or "september 30" in text:
        return True
    return False


def _is_total_row_label(label: str) -> bool:
    normalized = label.strip().lower()
    if not normalized:
        return False
    return normalized.startswith(
        (
            "total ",
            "net ",
            "gross profit",
            "operating profit",
            "profit for",
            "profit before",
            "cash and cash equivalents at end",
            "net cash provided",
            "net cash used",
            "total comprehensive income",
        )
    )


def _autofit_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            length = max(len(part) for part in lines)
            if length > max_len:
                max_len = length
        letter = ws.cell(row=1, column=col_idx).column_letter
        if col_idx == 1:
            width = max(36, min(70, max_len + 2))
        elif _is_notes_column(ws, col_idx):
            width = max(10, min(14, max_len + 2))
        else:
            width = max(14, min(20, max_len + 2))
        ws.column_dimensions[letter].width = width

    _auto_height_wrapped_rows(ws)


def _is_notes_column(ws, col_idx: int) -> bool:
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
        if value in {"note", "notes"}:
            return True
    return False


def _auto_height_wrapped_rows(ws) -> None:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            parts = text.splitlines() or [""]
            column_width = ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width or 10
            usable_width = max(8, int(column_width) - 2)
            explicit_lines = len(parts)
            approx_lines = max(
                1,
                max(
                    (len(part) + usable_width - 1) // usable_width
                    for part in parts
                ),
            )
            max_lines = max(max_lines, explicit_lines, approx_lines)
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = min(90, 15 * max_lines)
